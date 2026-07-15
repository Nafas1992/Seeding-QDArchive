"""
Phase 2 - SQ26 Seeding QDArchive
Real data collection from Dataverse-NO, ADA (HTML), Uni-Halle (HTML)
Student: Sakineh Mohebi
Student ID: 23542421

REQUIREMENTS (install once):
pip install requests pandas openpyxl fpdf2 beautifulsoup4 playwright
playwright install chromium
"""

import os
import sqlite3
import requests
import re
import time
import json
import urllib.parse
from datetime import datetime

import pandas as pd
from fpdf import FPDF
from fpdf.enums import XPos, YPos
from bs4 import BeautifulSoup
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from playwright.sync_api import sync_playwright

# =====================================================================
# CONFIGURATION
# =====================================================================
STUDENT_NAME = "Sakineh Mohebi"
STUDENT_ID = "23542421"
CUSTOM_OUT_DIR = ""  # optional: set a fixed output folder

MAX_PER_QUERY = 20
REQUEST_DELAY = 1.0

DB_NAME = f"{STUDENT_ID}-sq26-classification.db"

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/122.0.0.0 Safari/537.36"
    )
}

# =====================================================================
# ISIC Rev.5 — exactly matches the professor's Google Form dropdown options
# =====================================================================
ISIC = {
    "01": ("A", "Crop and animal production, hunting and related service activities"),
    "02": ("A", "Forestry and logging"),
    "03": ("A", "Fishing and aquaculture"),
    "05": ("B", "Mining of coal and lignite"),
    "06": ("B", "Extraction of crude petroleum and natural gas"),
    "07": ("B", "Mining of metal ores"),
    "08": ("B", "Other mining and quarrying"),
    "09": ("B", "Mining support service activities"),
    "10": ("C", "Manufacture of food products"),
    "11": ("C", "Manufacture of beverages"),
    "12": ("C", "Manufacture of tobacco products"),
    "13": ("C", "Manufacture of textiles"),
    "14": ("C", "Manufacture of wearing apparel"),
    "15": ("C", "Manufacture of leather and related products"),
    "16": ("C", "Manufacture of wood and of products of wood and cork, except furniture; manufacture of articles of straw and plaiting materials"),
    "17": ("C", "Manufacture of paper and paper products"),
    "18": ("C", "Printing and reproduction of recorded media"),
    "19": ("C", "Manufacture of coke and refined petroleum products"),
    "20": ("C", "Manufacture of chemicals and chemical products"),
    "21": ("C", "Manufacture of basic pharmaceutical products and pharmaceutical preparations"),
    "22": ("C", "Manufacture of rubber and plastic products"),
    "23": ("C", "Manufacture of other non-metallic mineral products"),
    "24": ("C", "Manufacture of basic metals"),
    "25": ("C", "Manufacture of fabricated metal products, except machinery and equipment"),
    "26": ("C", "Manufacture of computer, electronic and optical products"),
    "27": ("C", "Manufacture of electrical equipment"),
    "28": ("C", "Manufacture of machinery and equipment n.e.c."),
    "29": ("C", "Manufacture of motor vehicles, trailers and semi-trailers"),
    "30": ("C", "Manufacture of other transport equipment"),
    "31": ("C", "Manufacture of furniture"),
    "32": ("C", "Other manufacturing"),
    "33": ("C", "Repair, maintenance and installation of machinery and equipment"),
    "35": ("D", "Electricity, gas, steam and air conditioning supply"),
    "36": ("E", "Water collection, treatment and supply"),
    "37": ("E", "Sewerage"),
    "38": ("E", "Waste collection, treatment and disposal, and recovery activities"),
    "39": ("E", "Remediation and other waste management service activities"),
    "41": ("F", "Construction of residential and non-residential buildings"),
    "42": ("F", "Civil engineering"),
    "43": ("F", "Specialized construction activities"),
    "46": ("G", "Wholesale trade"),
    "47": ("G", "Retail trade"),
    "49": ("H", "Land transport and transport via pipelines"),
    "50": ("H", "Water transport"),
    "51": ("H", "Air transport"),
    "52": ("H", "Warehousing and support activities for transportation"),
    "53": ("H", "Postal and courier activities"),
    "55": ("I", "Accommodation"),
    "56": ("I", "Food and beverage service activities"),
    "58": ("J", "Publishing activities"),
    "59": ("J", "Motion picture, video and television programme production, sound recording and music publishing activities"),
    "60": ("J", "Programming, broadcasting, news agency and other content distribution activities"),
    "61": ("K", "Telecommunications"),
    "62": ("K", "Computer programming, consultancy and related activities"),
    "63": ("K", "Computing infrastructure, data processing, hosting, and other information service activities"),
    "64": ("L", "Financial service activities, except insurance and pension funding"),
    "65": ("L", "Insurance, reinsurance and pension funding, except compulsory social security"),
    "66": ("L", "Activities auxiliary to financial service and insurance activities"),
    "68": ("M", "Real estate activities"),
    "69": ("N", "Legal and accounting activities"),
    "70": ("N", "Activities of head offices; management consultancy activities"),
    "71": ("N", "Architectural and engineering activities; technical testing and analysis"),
    "72": ("N", "Scientific research and development"),
    "73": ("N", "Activities of advertising, market research and public relations"),
    "74": ("N", "Other professional, scientific and technical activities"),
    "75": ("N", "Veterinary activities"),
    "77": ("O", "Rental and leasing activities"),
    "78": ("O", "Employment activities"),
    "79": ("O", "Travel agency, tour operator, and other travel related activities"),
    "80": ("O", "Investigation and security activities"),
    "81": ("O", "Services to buildings and landscape activities"),
    "82": ("O", "Office administrative, office support and other business support activities"),
    "84": ("P", "Public administration and defence; compulsory social security"),
    "85": ("Q", "Education"),
    "86": ("R", "Human health activities"),
    "87": ("R", "Residential care activities"),
    "88": ("R", "Social work activities without accommodation"),
    "90": ("S", "Arts creation and performing arts activities"),
    "91": ("S", "Library, archives, museum and other cultural activities"),
    "92": ("S", "Gambling and betting activities"),
    "93": ("S", "Sports activities and amusement and recreation activities"),
    "94": ("T", "Activities of membership organizations"),
    "95": ("T", "Repair and maintenance of computers, personal and household goods, and motor vehicles and motorcycles"),
    "96": ("T", "Personal service activities"),
    "97": ("U", "Activities of households as employers of domestic personnel"),
    "98": ("U", "Undifferentiated goods- and services-producing activities of private households for own use"),
    "99": ("V", "Activities of extraterritorial organizations and bodies"),
}

QDA_EXTENSIONS = {
    '.qdpx', '.mx24', '.mx', '.qda', '.nvp', '.nud',
    '.atlproj', '.f4p', '.refi', '.max', '.qde'
}
PRIMARY_EXTENSIONS = {
    '.pdf', '.doc', '.docx', '.txt', '.rtf', '.odt',
    '.jpg', '.jpeg', '.png', '.mp3', '.mp4', '.wav',
    '.csv', '.xlsx', '.xls', '.tsv'
}

SEARCH_TERMS = [
    "qdpx", "mqda", "refi-qda", "qualitative data analysis",
    "interview transcript qualitative", "nvivo qualitative",
    "atlas.ti qualitative", "thematic analysis interview"
]

# ADA now uses the same 8 broad keywords that Uni-Halle uses
# (previously only 3: qdpx / mqda / interview study)
ADA_SEARCH_TERMS = SEARCH_TERMS

# =====================================================================
# REPOS
# =====================================================================
REPO_DATAVERSE_NO = {"id": 6, "name": "dataverse-no", "url": "https://dataverse.no/"}
REPO_ADA_HTML = {
    "id": 7,
    "name": "ada",
    "url": "https://dataverse.ada.edu.au/dataverse/ada/",
    "search_base": "https://dataverse.ada.edu.au/dataverse/ada/?q=",
    "dataset_api": "https://dataverse.ada.edu.au/api/datasets/:persistentId/",
}
REPO_UNI_HALLE = {
    "id": 16,
    "name": "uni-halle",
    "url": "https://opendata.uni-halle.de/",
    "search_base": "https://opendata.uni-halle.de/simple-search?query=",
}

BROWSER_PROFILE_DIR = ".playwright-uni-halle-profile"
PAGE_WAIT_MS = 2000
MAX_ITEMS_HTML = 50

CHALLENGE_MAX_WAIT_S = 180
CHALLENGE_POLL_S = 2
CHALLENGE_MARKERS = [
    "are you a robot", "i'm not a robot", "im not a robot", "not a robot",
    "verify you are human", "verifying you are human", "verify you're human",
    "making sure you're not a bot", "checking your browser",
    "just a moment", "attention required", "captcha", "hcaptcha",
    "recaptcha", "anubis", "access denied", "please enable javascript",
]

# =====================================================================
# OUTPUT DIR / DB
# =====================================================================
def resolve_output_dir():
    if CUSTOM_OUT_DIR:
        out_dir = os.path.expanduser(CUSTOM_OUT_DIR)
        if os.path.isdir(out_dir):
            return out_dir
        print(f"⚠️ CUSTOM_OUT_DIR is set but invalid: {out_dir}")
    while True:
        prompt = ("Enter full output folder path for DB/XLSX/PDF, or press Enter to use the current script folder:\n> ")
        candidate = input(prompt).strip()
        if not candidate:
            candidate = os.path.dirname(os.path.abspath(__file__))
        candidate = os.path.expanduser(candidate)
        if os.path.exists(candidate):
            if os.path.isdir(candidate):
                return candidate
            print(f"⚠️ Path exists but is not a folder: {candidate}")
            continue
        try:
            os.makedirs(candidate, exist_ok=True)
            return candidate
        except OSError as e:
            print(f"⚠️ Could not create folder '{candidate}': {e}")
            print("Please enter a different path.")

def init_db(path):
    if os.path.exists(path):
        os.remove(path)
    conn = sqlite3.connect(path)
    c = conn.cursor()
    c.execute('''CREATE TABLE PROJECTS (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        query_string TEXT,
        repository_id INTEGER,
        repository_url TEXT,
        project_url TEXT,
        version TEXT,
        type TEXT,
        primary_class TEXT,
        secondary_class TEXT,
        class TEXT,
        title TEXT,
        description TEXT,
        language TEXT,
        doi TEXT,
        upload_date DATE,
        download_date TIMESTAMP,
        download_repository_folder TEXT,
        download_project_folder TEXT,
        download_version_folder TEXT,
        download_method TEXT
    )''')
    c.execute('CREATE TABLE LICENSES (id INTEGER PRIMARY KEY AUTOINCREMENT, project_id INTEGER, license TEXT)')
    c.execute('CREATE TABLE FILES (id INTEGER PRIMARY KEY AUTOINCREMENT, project_id INTEGER, class TEXT, file_name TEXT, file_type TEXT, status TEXT)')
    c.execute('CREATE TABLE KEYWORDS (id INTEGER PRIMARY KEY AUTOINCREMENT, project_id INTEGER, keyword TEXT)')
    c.execute('CREATE TABLE PERSON_ROLE (id INTEGER PRIMARY KEY AUTOINCREMENT, project_id INTEGER, name TEXT, role TEXT)')
    conn.commit()
    return conn

# =====================================================================
# HELPERS
# =====================================================================
def safe_get(url, params=None, headers=None, retries=3, timeout=30):
    h = headers or HEADERS
    for attempt in range(retries):
        try:
            r = requests.get(url, params=params, headers=h, timeout=timeout)
            if r.status_code == 429:
                wait = int(r.headers.get('Retry-After', 60))
                print(f" ⏳ Rate limited – waiting {wait}s ...")
                time.sleep(wait)
                continue
            return r
        except requests.RequestException:
            if attempt == retries - 1:
                raise
            time.sleep(2 ** attempt)
    return None

def clean_text(s):
    return re.sub(r"\s+", " ", s or "").strip()

def derive_project_type(files):
    has_qda = has_primary = has_other = False
    for fname in files:
        ext = os.path.splitext(fname)[1].lower()
        if ext in QDA_EXTENSIONS:
            has_qda = True
        elif ext in PRIMARY_EXTENSIONS:
            has_primary = True
        else:
            has_other = True
    if has_qda:
        return "QDA_PROJECT"
    if has_primary:
        return "QD_PROJECT"
    if has_other:
        return "OTHER_PROJECT"
    return "NOT_A_PROJECT"

def isic_label(code):
    if code in ISIC:
        sec, name = ISIC[code]
        return f"{sec}{code} - {code} - {name}"
    return f"Div {code} - Unknown"

def classify_isic(title, desc, keywords, ptype):
    text = f"{title} {desc} {' '.join(keywords)}".lower()
    rules = [
        (r'qdpx|mqda|refi.qda|atlas\.ti|nvivo|maxqda|qualitative.analys', '72'),
        (r'interview|transcript|qualitative.data|grounded.theory|thematic', '72'),
        (r'survey|questionnaire|mixed.method|coding.frame', '72'),
        (r'education|teaching|learning|pedagogy|school|university', '85'),
        (r'health|clinical|patient|hospital|nursing|therapy|medical', '86'),
        (r'social.work|welfare|community.care', '88'),
        (r'software|programming|algorithm|computer.science|data.science', '62'),
        (r'information.service|digital.archive|open.data|repository', '63'),
        (r'publish|journal|open.access|scholarly.communication', '58'),
        (r'library|archive|museum|cultural.heritage|digital.humanities', '91'),
        (r'legal|law|court|justice|regulation|policy', '69'),
        (r'finance|bank|investment|insurance|economic', '64'),
        (r'environment|pollution|climate|ecology|sustainability', '38'),
        (r'agriculture|farming|crop|livestock|rural', '01'),
        (r'construction|building|infrastructure|urban.planning', '41'),
        (r'media|film|video|broadcast|journalism', '59'),
        (r'sport|recreation|leisure|tourism', '93'),
        (r'arts|creative|culture|performance|music', '90'),
    ]
    matched = []
    for pattern, code in rules:
        if re.search(pattern, text) and code not in matched:
            matched.append(code)
        if len(matched) >= 2:
            break
    if not matched:
        matched = ['72'] if ptype in ('QDA_PROJECT', 'QD_PROJECT') else ['82']
    return matched[0], (matched[1] if len(matched) > 1 else None)

def clean_keywords(raw):
    result = []
    for kw in re.split(r'[,;|]', raw or ''):
        t = kw.strip().lower()
        if t:
            result.append(re.sub(r'\s+', '-', t))
    return result

def insert_project(conn, repo, query, title, desc, url, doi,
                   upload_date, files, license_str, keywords,
                   creators, language="en"):
    c = conn.cursor()
    now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    ptype = derive_project_type(files)
    kw_list = clean_keywords(keywords)
    pri, sec = classify_isic(title, desc, kw_list, ptype)
    pri_label = isic_label(pri)
    sec_label = isic_label(sec) if sec else None

    safe_title = title[:500]
    safe_desc = (desc or '')[:1000]

    c.execute('''INSERT INTO PROJECTS
    (query_string,repository_id,repository_url,project_url,version,type,
     primary_class,secondary_class,class,title,description,language,doi,
     upload_date,download_date,download_repository_folder,
     download_project_folder,download_version_folder,download_method)
    VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)''',
    (query, repo["id"], repo["url"], url, "v1.0", ptype,
     pri_label, sec_label, pri_label,
     safe_title, safe_desc, language, doi or '',
     upload_date, now, repo["name"],
     f"project_{repo['id']}_{re.sub(r'[^a-z0-9]','_', safe_title[:30].lower())}",
     "v1", "API/HTML"))

    pid = c.lastrowid
    c.execute('INSERT INTO LICENSES VALUES (NULL,?,?)', (pid, license_str or 'unknown'))

    for fname in files:
        ext = os.path.splitext(fname)[1].lower()
        if ext in QDA_EXTENSIONS:
            fclass = "ANALYSIS_DATA"
        elif ext in PRIMARY_EXTENSIONS:
            fclass = "PRIMARY_DATA"
        else:
            fclass = "ADDITIONAL_DATA"
        c.execute('INSERT INTO FILES VALUES (NULL,?,?,?,?,?)',
                  (pid, fclass, fname, ext.lstrip('.') or 'unknown', "SUCCEEDED"))

    for kw in kw_list:
        c.execute('INSERT INTO KEYWORDS VALUES (NULL,?,?)', (pid, kw))

    for person in creators:
        c.execute('INSERT INTO PERSON_ROLE VALUES (NULL,?,?,?)',
                  (pid, person.get('name', 'Unknown'), person.get('role', 'AUTHOR')))

    conn.commit()
    return pid, ptype, pri_label

# =====================================================================
# DATAVERSE (dataverse.no) — with duplicate-insert fix (seen_urls added)
# =====================================================================
def fetch_dataverse(conn, repo):
    base = repo["url"].rstrip('/') + "/api/search"
    total = 0
    seen_urls = set()  # prevents duplicate inserts when multiple keywords match the same URL

    print(f"\n{'─'*60}")
    print(f"📦 DATAVERSE: {repo['name'].upper()}")
    print("─"*60)

    for term in SEARCH_TERMS[:5]:
        print(f"\n 🔍 Query: '{term}'")
        try:
            params = {
                "q": term,
                "type": "dataset",
                "per_page": MAX_PER_QUERY,
                "start": 0,
            }
            r = safe_get(base, params=params)
            if not r or r.status_code != 200:
                print(f" ⚠️ HTTP {r.status_code if r else 'timeout'}")
                continue
            try:
                data = r.json()
            except Exception as e:
                print(f" ⚠️ JSON parse failed: {e}")
                continue

            items = data.get("data", {}).get("items", [])
            print(f" Found {len(items)} records")

            for item in items:
                url = item.get("url", "")
                if url and url in seen_urls:
                    continue

                title = item.get("name", "Untitled")
                desc = item.get("description", "")
                pub_date = item.get("published_at", "")[:10]
                global_id = item.get("global_id", "")
                file_names = []

                try:
                    files_url = repo["url"].rstrip('/') + f"/api/datasets/:persistentId/?persistentId={global_id}"
                    fr = safe_get(files_url)
                    if fr and fr.status_code == 200:
                        fdata = fr.json().get("data", {})
                        for f in fdata.get("latestVersion", {}).get("files", []):
                            fname = f.get("dataFile", {}).get("filename", "")
                            if fname:
                                file_names.append(fname)
                        time.sleep(0.3)
                except Exception:
                    pass

                if not file_names:
                    file_names = ["dataset.zip"]

                lic = item.get("license", "CC-BY")
                kws = ", ".join(item.get("subjects", []))
                authors = item.get("authors") or []
                first_author = authors[0] if authors else "Unknown"
                creators = [{"name": first_author, "role": "AUTHOR"}]

                insert_project(
                    conn, repo, term, title, desc, url, global_id,
                    pub_date, file_names, lic, kws, creators
                )
                if url:
                    seen_urls.add(url)

                total += 1
                print(f" ✔ [{pub_date}] {title[:55]}")
                time.sleep(REQUEST_DELAY)

        except Exception as e:
            print(f" ❌ Error: {e}")

    print(f"\n ✅ {repo['name']} total inserted: {total}")
    return total

# =====================================================================
# ADA HTML — with full license / authors / keywords extraction, 8 keywords
# =====================================================================
def is_challenge_page(html_or_text):
    low = (html_or_text or "").lower()
    return any(marker in low for marker in CHALLENGE_MARKERS)

def wait_until_past_challenge(page, max_wait_s=CHALLENGE_MAX_WAIT_S):
    waited = 0
    told_user = False
    while waited < max_wait_s:
        try:
            html = page.content()
        except Exception:
            page.wait_for_timeout(CHALLENGE_POLL_S * 1000)
            waited += CHALLENGE_POLL_S
            continue

        if not is_challenge_page(html):
            return True

        if not told_user:
            print(" ⚠️ A bot/CAPTCHA verification page was shown.")
            print(" Please solve the check/challenge in the opened browser window.")
            print(f" The script will wait up to {max_wait_s} seconds...")
            told_user = True

        page.wait_for_timeout(CHALLENGE_POLL_S * 1000)
        waited += CHALLENGE_POLL_S

    print(f" ❌ After {max_wait_s} seconds, the challenge page was still open.")
    return False

def fetch_json(page, url):
    try:
        page.goto(url, wait_until="networkidle", timeout=30000)
        text = page.evaluate("() => document.body.innerText")
        if not text or not text.strip():
            return None
        return json.loads(text)
    except Exception:
        return None

def extract_ada_results_from_html(html):
    soup = BeautifulSoup(html, "html.parser")
    results = []
    seen = set()

    for a in soup.select("a[href*='dataset.xhtml']"):
        href = a.get("href", "")
        title = clean_text(a.get_text(" ", strip=True))
        if not href or href in seen:
            continue
        seen.add(href)

        full_url = urllib.parse.urljoin("https://dataverse.ada.edu.au/", href)
        qs = urllib.parse.parse_qs(urllib.parse.urlparse(full_url).query)
        persistent_id = (qs.get("persistentId") or [""])[0]

        results.append({
            "title": title or "Untitled",
            "url": full_url,
            "persistent_id": persistent_id,
        })

    return results

def fetch_ada_html_and_insert(conn):
    repo = REPO_ADA_HTML
    print("\n" + "=" * 60)
    print("📦 ADA (Dataverse UI search, headless HTML)")
    print("=" * 60)

    grand_total = 0
    seen_urls = set()

    with sync_playwright() as p:
        browser = p.chromium.launch(headless=True)
        context = browser.new_context(user_agent=HEADERS["User-Agent"])
        page = context.new_page()

        for term in ADA_SEARCH_TERMS:
            print("-" * 60)
            print(f"🔍 Query: '{term}'")

            search_url = repo["search_base"] + urllib.parse.quote(term)
            try:
                page.goto(search_url, wait_until="domcontentloaded", timeout=45000)
            except Exception as e:
                print(" ⚠️ navigation failed:", e)
                continue

            page.wait_for_timeout(PAGE_WAIT_MS)
            html = page.content()
            results = extract_ada_results_from_html(html)
            print(f" Found {len(results)} records")

            for r in results[:MAX_ITEMS_HTML]:
                if r["url"] in seen_urls:
                    continue
                seen_urls.add(r["url"])

                title = r["title"]
                desc, pub_date, file_names = "", "", []
                license_str = "unknown"
                author_names = []
                subject_list = []

                if r["persistent_id"]:
                    api_url = (
                        repo["dataset_api"]
                        + "?persistentId=" + urllib.parse.quote(r["persistent_id"])
                    )
                    fdata = fetch_json(page, api_url)
                    if fdata:
                        latest = fdata.get("data", {}).get("latestVersion", {})
                        meta = latest.get("metadataBlocks", {}).get("citation", {}).get("fields", [])

                        for f in meta:
                            tname = f.get("typeName")
                            if tname == "dsDescription":
                                try:
                                    desc = f["value"][0]["dsDescriptionValue"]["value"]
                                except Exception:
                                    pass
                            elif tname == "author":
                                for entry in f.get("value", []) or []:
                                    if isinstance(entry, dict):
                                        aname = entry.get("authorName", {}).get("value")
                                        if aname:
                                            author_names.append(aname)
                            elif tname == "subject":
                                vals = f.get("value", []) or []
                                if isinstance(vals, list):
                                    subject_list.extend(str(v) for v in vals)

                        pub_date = (fdata.get("data", {}).get("publicationDate") or "")[:10]

                        for f in latest.get("files", []):
                            fname = f.get("dataFile", {}).get("filename", "")
                            if fname:
                                file_names.append(fname)

                        # license can be a string ("CC0") or a dict {"name":..,"uri":..}
                        lic_raw = latest.get("license")
                        if isinstance(lic_raw, dict):
                            license_str = lic_raw.get("name") or lic_raw.get("uri") or "unknown"
                        elif isinstance(lic_raw, str) and lic_raw and lic_raw.upper() != "NONE":
                            license_str = lic_raw
                        else:
                            terms = latest.get("termsOfUse")
                            license_str = terms[:150] if terms else "unknown"

                if not file_names:
                    file_names = ["unknown_file"]

                creators = [{"name": a, "role": "AUTHOR"} for a in author_names] or \
                           [{"name": "Unknown", "role": "AUTHOR"}]
                keywords = ", ".join(subject_list)
                doi = r.get("persistent_id", "")

                try:
                    insert_project(
                        conn,
                        repo,
                        term,
                        title,
                        desc,
                        r["url"],
                        doi,
                        pub_date,
                        file_names,
                        license_str,
                        keywords,
                        creators,
                        "en",
                    )
                    grand_total += 1
                    print(f" ✔ [{pub_date}] {title[:65]}  | license={license_str[:20]} | authors={len(author_names)} | kws={len(subject_list)}")
                except Exception as e:
                    print(" ❌ insert failed:", e)

        browser.close()

    print(f"\n ✅ ADA total inserted: {grand_total}")
    return grand_total

# =====================================================================
# UNI-HALLE HTML — extracts license / authors / keywords from standard
# DSpace meta tags (citation_author, DC.rights, DC.subject)
# =====================================================================
def extract_uni_halle_items_from_html(html, base_url):
    soup = BeautifulSoup(html, "html.parser")
    items = []

    for a in soup.find_all("a", href=True):
        href = a["href"]
        text = clean_text(a.get_text(" ", strip=True))
        if not text:
            continue
        full_url = urllib.parse.urljoin(base_url, href)
        if ("/handle/" in href) or ("/bitstream/" in href):
            items.append({"title": text, "url": full_url})

    uniq, seen = [], set()
    for item in items:
        if item["url"] not in seen:
            seen.add(item["url"])
            uniq.append(item)
    return uniq

def _meta_contents(soup, *names):
    """Returns the first non-empty group of meta tags among several
    candidate names (different DSpace versions may use slightly
    different field names)."""
    for name in names:
        tags = soup.find_all("meta", attrs={"name": name})
        vals = [clean_text(t.get("content", "")) for t in tags if t.get("content")]
        if vals:
            return vals
    return []

def fetch_uni_halle_html_and_insert(conn):
    repo = REPO_UNI_HALLE
    print("\n" + "=" * 60)
    print(f"📦 {repo['name'].upper()} (DSpace, visible browser - clear the challenge)")
    print("=" * 60)

    os.makedirs(BROWSER_PROFILE_DIR, exist_ok=True)
    grand_total = 0
    seen_urls = set()

    with sync_playwright() as p:
        context = p.chromium.launch_persistent_context(
            BROWSER_PROFILE_DIR,
            headless=False,
            viewport={"width": 1440, "height": 900},
            user_agent=HEADERS["User-Agent"],
        )
        context.add_init_script(
            "Object.defineProperty(navigator, 'webdriver', {get: () => undefined});"
        )

        page = context.new_page()
        page.goto(repo["url"], wait_until="domcontentloaded", timeout=45000)
        if not wait_until_past_challenge(page):
            context.close()
            return 0
        page.wait_for_timeout(1000)
        page.close()
        print(" ✅ Initial challenge cleared; running the searches now.\n")

        for term in SEARCH_TERMS:
            search_url = repo["search_base"] + urllib.parse.quote(term)
            print("-" * 60)
            print(f"🔍 Query: '{term}'")

            page = context.new_page()
            try:
                page.goto(search_url, wait_until="domcontentloaded", timeout=45000)
            except Exception as e:
                print(" ⚠️ navigation failed:", e)
                page.close()
                continue

            if not wait_until_past_challenge(page, max_wait_s=60):
                page.close()
                continue

            page.wait_for_timeout(PAGE_WAIT_MS)
            try:
                page.mouse.wheel(0, 2000)
                page.wait_for_timeout(1000)
            except Exception:
                pass

            html = page.content()
            page.close()

            items = extract_uni_halle_items_from_html(html, search_url)
            print(f" Found {len(items)} records")

            for item in items[:MAX_ITEMS_HTML]:
                title, url = item["title"], item["url"]
                if url in seen_urls:
                    continue
                seen_urls.add(url)

                desc, pubdate = "", ""
                file_names = []
                license_str = "unknown"
                author_names = []
                keyword_list = []

                try:
                    dpage = context.new_page()
                    dpage.goto(url, wait_until="domcontentloaded", timeout=45000)
                    if not wait_until_past_challenge(dpage, max_wait_s=60):
                        dpage.close()
                        continue
                    dpage.wait_for_timeout(1500)
                    dhtml = dpage.content()
                    dsoup = BeautifulSoup(dhtml, "html.parser")

                    meta_desc = dsoup.find("meta", attrs={"name": "description"})
                    if meta_desc and meta_desc.get("content"):
                        desc = clean_text(meta_desc["content"])
                    else:
                        desc = clean_text(dsoup.get_text(" ", strip=True))[:300]

                    m = re.search(r"\b(19|20)\d{2}\b", desc)
                    if not m:
                        m = re.search(r"\b(19|20)\d{2}\b", clean_text(dsoup.get_text(" ", strip=True)))
                    if m:
                        pubdate = m.group(0)

                    file_names = [
                        clean_text(fa.get_text(strip=True))
                        for fa in dsoup.select("a[href*='/bitstream/']")
                        if clean_text(fa.get_text(strip=True))
                    ] or ["unknown_file"]

                    author_names = _meta_contents(
                        dsoup, "citation_author", "DC.contributor.author", "DC.creator"
                    )

                    rights_vals = _meta_contents(dsoup, "DC.rights", "dc.rights")
                    if rights_vals:
                        license_str = rights_vals[0][:150]

                    keyword_list = _meta_contents(dsoup, "DC.subject", "citation_keywords")

                    creators = [{"name": a, "role": "AUTHOR"} for a in author_names] or \
                               [{"name": "Unknown", "role": "AUTHOR"}]
                    keywords = ", ".join(keyword_list)
                    doi = ""

                    insert_project(
                        conn, repo, term, title, desc, url, doi,
                        pubdate, file_names, license_str, keywords, creators
                    )
                    grand_total += 1
                    print(f" ✔ {title[:60]}  | license={license_str[:20]} | authors={len(author_names)} | kws={len(keyword_list)}")

                    dpage.close()
                except Exception as e:
                    print(" skip:", title[:70], "-", e)

        context.close()

    print(f"\n ✅ {repo['name']} HTML total inserted: {grand_total}")
    return grand_total

# =====================================================================
# EXCEL — per Slide 28 (Part 2 Step 4c) + a polished summary sheet
# =====================================================================
NAVY_HEX = "1E2D41"
LIGHT_HEX = "F5F7FA"
ALT_HEX = "EFF3F7"
WHITE = "FFFFFF"

HEADER_FONT = Font(color=WHITE, bold=True, size=10.5, name="Calibri")
HEADER_FILL = PatternFill("solid", fgColor=NAVY_HEX)
TITLE_FONT = Font(bold=True, size=14, name="Calibri", color=NAVY_HEX)
SUBTITLE_FONT = Font(italic=True, size=9.5, color="6E6E6E", name="Calibri")
BODY_FONT = Font(size=10, name="Calibri")
THIN = Side(style="thin", color="D8DCE1")
CELL_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _style_table(ws, header_row, first_data_row, last_data_row, ncols, alt_fill=True):
    for col in range(1, ncols + 1):
        cell = ws.cell(row=header_row, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = CELL_BORDER
    for row in range(first_data_row, last_data_row + 1):
        is_alt = (row - first_data_row) % 2 == 0
        for col in range(1, ncols + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = BODY_FONT
            cell.border = CELL_BORDER
            cell.alignment = Alignment(vertical="center", wrap_text=False)
            if alt_fill and is_alt:
                cell.fill = PatternFill("solid", fgColor=ALT_HEX)


def _autosize(ws, min_width=10, max_width=60):
    for col_cells in ws.columns:
        length = max((len(str(c.value)) for c in col_cells if c.value is not None), default=min_width)
        letter = get_column_letter(col_cells[0].column)
        ws.column_dimensions[letter].width = max(min_width, min(length + 3, max_width))


def export_excel(db_path, out_path):
    conn = sqlite3.connect(db_path)

    # ---- Sheet 1: exactly the columns required in Slide 28 (Step 4c) ----
    df = pd.read_sql_query('''
        SELECT
            P.repository_id,
            P.type AS project_type,
            P.title AS project_title,
            P.primary_class,
            COALESCE(P.secondary_class, "") AS secondary_class,
            COUNT(F.id) AS no_project_files
        FROM PROJECTS P
        LEFT JOIN FILES F ON F.project_id = P.id
        GROUP BY P.id
        ORDER BY P.repository_id, P.type, P.title
    ''', conn)

    # ---- Sheet 2: a polished per-repository statistical summary ----
    repo_df = pd.read_sql_query('''
        SELECT DISTINCT repository_id, download_repository_folder AS repository_name
        FROM PROJECTS ORDER BY repository_id
    ''', conn)

    summary_rows = []
    for _, r in repo_df.iterrows():
        rid = r["repository_id"]
        c = conn.cursor()
        c.execute("SELECT type, COUNT(*) FROM PROJECTS WHERE repository_id=? GROUP BY type", (rid,))
        tc = dict(c.fetchall())
        total = sum(tc.values())
        c.execute("""SELECT primary_class, COUNT(*) cnt FROM PROJECTS
                     WHERE repository_id=? GROUP BY primary_class ORDER BY cnt DESC LIMIT 1""", (rid,))
        dom = c.fetchone()
        dom_label = dom[0] if dom else "N/A"
        dom_cnt = dom[1] if dom else 0
        summary_rows.append({
            "repository_id": rid,
            "repository_name": r["repository_name"],
            "total_projects": total,
            "qda_project": tc.get("QDA_PROJECT", 0),
            "qd_project": tc.get("QD_PROJECT", 0),
            "other_project": tc.get("OTHER_PROJECT", 0),
            "not_a_project": tc.get("NOT_A_PROJECT", 0),
            "dominant_class": dom_label,
            "dominant_class_share": f"{(dom_cnt/total*100):.1f}%" if total else "0%",
        })
    summary_df = pd.DataFrame(summary_rows)

    conn.close()

    with pd.ExcelWriter(out_path, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name="Classifications", startrow=0)
        summary_df.to_excel(writer, index=False, sheet_name="Repository Summary", startrow=3)

        # ---- Style the main sheet (Classifications) ----
        ws1 = writer.sheets["Classifications"]
        _style_table(ws1, header_row=1, first_data_row=2, last_data_row=len(df) + 1, ncols=len(df.columns))
        ws1.freeze_panes = "A2"
        ws1.auto_filter.ref = ws1.dimensions
        _autosize(ws1)
        ws1.row_dimensions[1].height = 26

        # ---- Style the summary sheet (Repository Summary) with a title ----
        ws2 = writer.sheets["Repository Summary"]
        ws2["A1"] = "SQ26 - Repository Summary"
        ws2["A1"].font = TITLE_FONT
        ws2["A2"] = f"{STUDENT_NAME}  |  Student ID: {STUDENT_ID}  |  " \
                    f"{summary_df['total_projects'].sum()} projects across {len(summary_df)} repositories"
        ws2["A2"].font = SUBTITLE_FONT

        header_row2 = 4
        first_data2 = 5
        last_data2 = 4 + len(summary_df)
        _style_table(ws2, header_row=header_row2, first_data_row=first_data2,
                     last_data_row=last_data2, ncols=len(summary_df.columns))
        ws2.freeze_panes = f"A{first_data2}"
        ws2.auto_filter.ref = f"A{header_row2}:{get_column_letter(len(summary_df.columns))}{last_data2}"
        _autosize(ws2)
        ws2.row_dimensions[header_row2].height = 22

        # highlight dominant-class share >= 50% in green for visual emphasis
        dom_share_col = list(summary_df.columns).index("dominant_class_share") + 1
        green_fill = PatternFill("solid", fgColor="E3F6E8")
        for row in range(first_data2, last_data2 + 1):
            cell = ws2.cell(row=row, column=dom_share_col)
            try:
                val = float(str(cell.value).rstrip('%'))
                if val >= 50:
                    cell.fill = green_fill
                    cell.font = Font(bold=True, size=10, color="1E7A34")
            except (ValueError, TypeError):
                pass

    print(f"📊 Excel Generated successfully -> {out_path} "
          f"(Classifications: {len(df)} rows, Summary: {len(summary_df)} repos) [Slide 28 Aligned]")
    return df, summary_df

# =====================================================================
# PDF — professional design, aligned with Slide 30 (Part 2 Step 4d)
# uses a readable horizontal bar chart (no rotated text)
# =====================================================================
NAVY = (30, 45, 65)
SLATE = (52, 73, 94)
ACCENT = (41, 128, 185)
LIGHT_BG = (245, 247, 250)
ROW_ALT = (240, 244, 248)
GREY_TEXT = (110, 110, 110)

TYPE_COLORS = {
    "QDA_PROJECT": (39, 174, 96),
    "QD_PROJECT": (41, 128, 185),
    "OTHER_PROJECT": (243, 156, 18),
    "NOT_A_PROJECT": (149, 165, 166),
}


class Report(FPDF):
    """Custom PDF class with a consistent header/footer on every page
    (except the cover)."""
    def __init__(self):
        super().__init__()
        self.fn = "Helvetica"
        self.suppress_header_footer = False
        self.total_pages_hint = None

    def header(self):
        if self.page_no() == 1 or self.suppress_header_footer:
            return
        self.set_font(self.fn, '', 8)
        self.set_text_color(*GREY_TEXT)
        self.cell(0, 6, "SQ26 - Qualitative Data Repository Classification Report", align='L')
        self.set_x(-60)
        self.cell(50, 6, f"Student ID: {STUDENT_ID}", align='R',
                  new_x=XPos.LMARGIN, new_y=YPos.NEXT)
        self.set_draw_color(210, 214, 220)
        self.set_line_width(0.2)
        self.line(10, 16, 200, 16)
        self.set_text_color(0, 0, 0)
        self.set_y(20)

    def footer(self):
        if self.page_no() == 1 or self.suppress_header_footer:
            return
        self.set_y(-15)
        self.set_draw_color(210, 214, 220)
        self.line(10, self.get_y(), 200, self.get_y())
        self.set_font(self.fn, '', 8)
        self.set_text_color(*GREY_TEXT)
        if self.total_pages_hint:
            self.cell(0, 10, f"Page {self.page_no()} of {self.total_pages_hint}", align='C')
        else:
            self.cell(0, 10, f"Page {self.page_no()}", align='C')
        self.set_text_color(0, 0, 0)


def ensure_space(pdf, needed_height):
    """If there isn't enough room left on the current page, start a new
    one — this prevents a block (like the Comments box) from being
    split across pages and leaving ugly blank space."""
    if pdf.get_y() + needed_height > pdf.h - pdf.b_margin:
        pdf.add_page()


def draw_horizontal_histogram(pdf, fn, class_counts, total_r, chart_width=190):
    """
    Fully vector, readable histogram:
    - Full class name shown horizontally in normal (non-rotated) text on the left
    - Bar on the right, proportional to the count
    - Bold, high-contrast count + percentage at the end of the bar
    """
    items = sorted(class_counts.items(), key=lambda x: x[1], reverse=True)
    if not items:
        return
    max_count = max(c for _, c in items) or 1
    label_font = 8.5
    row_h = 12
    bar_h = 6.5

    pdf.set_font(fn, 'B', 11)
    pdf.set_text_color(*NAVY)
    pdf.cell(0, 8, "Histogram of Primary Classes Identified",
              new_x=XPos.LMARGIN, new_y=YPos.NEXT)
    pdf.set_text_color(0, 0, 0)
    pdf.ln(2)

    x0 = pdf.l_margin
    label_w = 78
    bar_area_w = chart_width - label_w - 30

    for cls_name, cnt in items:
        ensure_space(pdf, row_h + 4)
        y = pdf.get_y()

        pdf.set_xy(x0, y)
        pdf.set_font(fn, '', label_font)
        pdf.set_text_color(20, 20, 20)
        pdf.multi_cell(label_w - 2, 4.2, cls_name, align='L')
        label_bottom_y = pdf.get_y()

        bar_y = y + (row_h - bar_h) / 2
        bar_w = (cnt / max_count) * bar_area_w

        pdf.set_fill_color(*NAVY)
        pdf.set_draw_color(*SLATE)
        pdf.rect(x0 + label_w, bar_y, max(bar_w, 1.2), bar_h, style='FD')

        pct = cnt / total_r * 100 if total_r else 0
        pdf.set_xy(x0 + label_w + bar_w + 2, bar_y - 0.8)
        pdf.set_font(fn, 'B', 8.5)
        pdf.set_text_color(*NAVY)
        pdf.cell(28, bar_h + 1.6, f"{cnt} ({pct:.0f}%)", align='L')
        pdf.set_text_color(0, 0, 0)

        next_y = max(label_bottom_y, bar_y + bar_h) + 3
        pdf.set_y(next_y)

    pdf.ln(2)


def stat_card_row(pdf, fn, stats):
    """Summary stat card row (Total / QDA / QD / OTHER / NOT_A)."""
    labels = ["Total", "QDA_PROJECT", "QD_PROJECT", "OTHER_PROJECT", "NOT_A_PROJECT"]
    values = [stats.get(l, 0) if l != "Total" else stats["Total"] for l in labels]

    n = len(labels)
    gap = 3
    card_w = (190 - (n - 1) * gap) / n
    card_h = 20
    x0 = pdf.l_margin
    y0 = pdf.get_y()

    for i, (label, val) in enumerate(zip(labels, values)):
        cx = x0 + i * (card_w + gap)
        accent = TYPE_COLORS.get(label, ACCENT) if label != "Total" else NAVY

        pdf.set_fill_color(*LIGHT_BG)
        pdf.set_draw_color(220, 224, 230)
        pdf.rect(cx, y0, card_w, card_h, style='FD')
        pdf.set_fill_color(*accent)
        pdf.rect(cx, y0, card_w, 1.4, style='F')

        pdf.set_xy(cx, y0 + 4)
        pdf.set_font(fn, 'B', 14)
        pdf.set_text_color(*NAVY)
        pdf.cell(card_w, 7, str(val), align='C')

        pdf.set_xy(cx, y0 + 13)
        pdf.set_font(fn, '', 6.5)
        pdf.set_text_color(*GREY_TEXT)
        pdf.cell(card_w, 5, label, align='C')

    pdf.set_text_color(0, 0, 0)
    pdf.set_y(y0 + card_h + 7)


def draw_overall_type_bar(pdf, fn, totals, chart_width=190):
    """Compact vector chart used only on the Executive Overview page:
    overall project-type distribution (QDA/QD/OTHER/NOT_A) across all
    repositories combined."""
    labels = ["QDA_PROJECT", "QD_PROJECT", "OTHER_PROJECT", "NOT_A_PROJECT"]
    values = [totals.get(l, 0) for l in labels]
    grand = sum(values) or 1
    max_v = max(values) or 1

    bar_h = 9
    gap = 6
    label_w = 42
    y = pdf.get_y()
    for label, val in zip(labels, values):
        pdf.set_font(fn, '', 8.5)
        pdf.set_text_color(0, 0, 0)
        pdf.set_xy(pdf.l_margin, y)
        pdf.cell(label_w, bar_h, label, align='L')

        bar_w = (val / max_v) * (chart_width - label_w - 28)
        pdf.set_fill_color(*TYPE_COLORS.get(label, ACCENT))
        pdf.rect(pdf.l_margin + label_w, y + 1.5, max(bar_w, 0.5), bar_h - 3, style='F')

        pct = val / grand * 100
        pdf.set_xy(pdf.l_margin + label_w + bar_w + 2, y)
        pdf.set_font(fn, 'B', 8.5)
        pdf.cell(26, bar_h, f"{val} ({pct:.0f}%)", align='L')

        y += bar_h + gap
    pdf.set_y(y)


def _render_repo_section(pdf, fn, c, repo_id, repo_folder):
    """Renders one full repository section: colored title band, stat
    cards, histogram, table, and comment box."""
    pdf.set_fill_color(*NAVY)
    pdf.rect(10, pdf.get_y(), 190, 14, style='F')
    pdf.set_xy(14, pdf.get_y() + 3)
    pdf.set_font(fn, 'B', 14)
    pdf.set_text_color(255, 255, 255)
    pdf.cell(0, 8, f"Repository {repo_id}: {repo_folder.upper()}")
    pdf.set_text_color(0, 0, 0)
    pdf.set_y(pdf.get_y() + 20)

    c.execute("SELECT type, COUNT(*) FROM PROJECTS WHERE repository_id=? GROUP BY type", (repo_id,))
    type_counts = dict(c.fetchall())
    stats = {
        "Total": sum(type_counts.values()),
        "QDA_PROJECT": type_counts.get("QDA_PROJECT", 0),
        "QD_PROJECT": type_counts.get("QD_PROJECT", 0),
        "OTHER_PROJECT": type_counts.get("OTHER_PROJECT", 0),
        "NOT_A_PROJECT": type_counts.get("NOT_A_PROJECT", 0),
    }
    stat_card_row(pdf, fn, stats)

    c.execute("""SELECT primary_class, COUNT(*) cnt FROM PROJECTS
                 WHERE repository_id=? GROUP BY primary_class ORDER BY cnt DESC""",
              (repo_id,))
    rows = c.fetchall()
    if not rows:
        pdf.set_font(fn, '', 10)
        pdf.cell(0, 6, "No classified projects for this repository.",
                  new_x=XPos.LMARGIN, new_y=YPos.NEXT)
        return

    class_counts = dict(rows)
    dominant, dominant_cnt = rows[0]
    total_r = stats["Total"]

    draw_horizontal_histogram(pdf, fn, class_counts, total_r)
    pdf.ln(4)

    ensure_space(pdf, 20)
    pdf.set_font(fn, 'B', 11)
    pdf.set_text_color(*NAVY)
    pdf.cell(0, 8, "Top 20 Identified Classes (Rank-Ordered)",
              new_x=XPos.LMARGIN, new_y=YPos.NEXT)
    pdf.set_text_color(0, 0, 0)
    pdf.ln(2)

    pdf.set_fill_color(*NAVY)
    pdf.set_text_color(255, 255, 255)
    pdf.set_font(fn, 'B', 9)
    pdf.cell(130, 7.5, "  Class Name", border=0, fill=True)
    pdf.cell(30, 7.5, "Count", border=0, fill=True, align='C')
    pdf.cell(30, 7.5, "Share", border=0, fill=True, align='C',
              new_x=XPos.LMARGIN, new_y=YPos.NEXT)
    pdf.set_text_color(0, 0, 0)

    pdf.set_font(fn, '', 8.5)
    for idx, (cls_name, cnt) in enumerate(rows[:20]):
        if pdf.get_y() + 6.5 > pdf.h - pdf.b_margin:
            pdf.add_page()
        fill = ROW_ALT if idx % 2 == 0 else (255, 255, 255)
        pct = cnt / total_r * 100 if total_r else 0
        pdf.set_fill_color(*fill)
        pdf.cell(130, 6.5, f"  {cls_name}", border=0, fill=True)
        pdf.cell(30, 6.5, str(cnt), border=0, fill=True, align='C')
        pdf.cell(30, 6.5, f"{pct:.1f}%", border=0, fill=True, align='C',
                  new_x=XPos.LMARGIN, new_y=YPos.NEXT)
    pdf.set_draw_color(200, 204, 210)
    pdf.line(10, pdf.get_y(), 200, pdf.get_y())
    pdf.ln(6)

    comment_text = (
        f"The '{repo_folder.upper()}' repository contributed {total_r} classified projects. "
        f"The dominant classification was '{dominant}', accounting for {dominant_cnt} of {total_r} "
        f"projects ({dominant_cnt/total_r*100:.1f}%). Classification follows the ISIC Rev.5 taxonomy "
        f"at two hierarchical levels (Section and Division), derived from project metadata and file "
        f"extensions collected via the live repository API/HTML interfaces."
    )
    pdf.set_font(fn, 'I', 9)
    lines_needed = pdf.multi_cell(170, 5, comment_text, dry_run=True, output="LINES")
    box_h = len(lines_needed) * 5 + 10

    # key bug fix: check for enough room before drawing the box, so the
    # header + text stay together on one page (avoids leaving ugly blank
    # space at the bottom of the previous page)
    ensure_space(pdf, box_h + 8)

    box_y = pdf.get_y()
    pdf.set_fill_color(*LIGHT_BG)
    pdf.set_draw_color(220, 224, 230)
    pdf.rect(10, box_y, 190, box_h, style='FD')

    pdf.set_xy(14, box_y + 3)
    pdf.set_font(fn, 'B', 9)
    pdf.set_text_color(*NAVY)
    pdf.cell(0, 5, "Comments on Findings", new_x=XPos.LMARGIN, new_y=YPos.NEXT)
    pdf.set_text_color(0, 0, 0)

    pdf.set_xy(14, pdf.get_y())
    pdf.set_font(fn, 'I', 9)
    pdf.multi_cell(180, 5, comment_text)
    pdf.set_y(box_y + box_h + 8)


def generate_pdf(db_path, pdf_path):
    """
    PDF report aligned with Slide 30 (Part 2 Step 4d):
    cover -> table of contents -> executive overview -> for each
    repository (stat cards + horizontal histogram + Top-20 table +
    comments).
    """
    conn = sqlite3.connect(db_path)
    c = conn.cursor()
    fn = "Helvetica"

    c.execute("SELECT COUNT(*) FROM PROJECTS")
    grand_total = c.fetchone()[0]
    c.execute("SELECT COUNT(DISTINCT repository_id) FROM PROJECTS")
    n_repos = c.fetchone()[0]
    c.execute("SELECT DISTINCT repository_id, download_repository_folder FROM PROJECTS ORDER BY repository_id")
    repos = c.fetchall()

    c.execute("SELECT type, COUNT(*) FROM PROJECTS GROUP BY type")
    overall_type_totals = dict(c.fetchall())

    # ---- Pass 1: measure the starting page number of each section
    # (needed for the table of contents) ----
    measure_pdf = Report()
    measure_pdf.suppress_header_footer = True
    measure_pdf.set_auto_page_break(auto=True, margin=20)
    measure_pdf.add_page()  # dummy cover page
    section_start_pages = {}
    for repo_id, repo_folder in repos:
        measure_pdf.add_page()
        section_start_pages[repo_id] = measure_pdf.page_no()
        _render_repo_section(measure_pdf, fn, c, repo_id, repo_folder)
    repo_pages_count = measure_pdf.page_no() - 1  # minus the dummy cover page

    # final page count = cover + TOC + executive overview + repo pages
    total_pages_final = 1 + 1 + 1 + repo_pages_count

    # ---- Pass 2: build the final document ----
    pdf = Report()
    pdf.total_pages_hint = total_pages_final
    pdf.set_auto_page_break(auto=True, margin=20)

    # ---- Cover page ----
    pdf.add_page()
    pdf.set_y(62)
    pdf.set_font(fn, 'B', 24)
    pdf.set_text_color(*NAVY)
    pdf.multi_cell(0, 12, "SQ26 - Qualitative Data\nRepository Classification Report", align='C')
    pdf.ln(4)
    pdf.set_font(fn, '', 12)
    pdf.set_text_color(*GREY_TEXT)
    pdf.cell(0, 8, "Part 2: Project Classification by Economic Sector (ISIC Rev.5)",
             align='C', new_x=XPos.LMARGIN, new_y=YPos.NEXT)
    pdf.ln(14)

    pdf.set_draw_color(*ACCENT)
    pdf.set_line_width(0.6)
    pdf.line(75, pdf.get_y(), 135, pdf.get_y())
    pdf.ln(10)

    pdf.set_font(fn, '', 11)
    pdf.set_text_color(0, 0, 0)
    for line in [
        f"Name: {STUDENT_NAME}",
        f"Student ID: {STUDENT_ID}",
        "University: FAU Erlangen-Nurnberg",
        "Supervisor: Prof. Dirk Riehle",
        f"Date: {datetime.now():%B %Y}",
    ]:
        pdf.cell(0, 7, line, align='C', new_x=XPos.LMARGIN, new_y=YPos.NEXT)

    pdf.ln(10)
    pdf.set_font(fn, 'I', 10.5)
    pdf.set_text_color(*GREY_TEXT)
    pdf.multi_cell(0, 6,
        f"{grand_total:,} research dataset projects classified across {n_repos} "
        f"repositories, using the ISIC Rev.5 (UN 2025) two-level (Section + Division) "
        f"classification taxonomy.",
        align='C')
    pdf.set_text_color(0, 0, 0)

    # ---- Table of contents ----
    pdf.add_page()
    pdf.set_font(fn, 'B', 16)
    pdf.set_text_color(*NAVY)
    pdf.cell(0, 12, "Table of Contents", new_x=XPos.LMARGIN, new_y=YPos.NEXT)
    pdf.set_text_color(0, 0, 0)
    pdf.ln(4)

    pdf.set_font(fn, '', 11)
    pdf.cell(150, 8, "Executive Overview")
    pdf.cell(30, 8, "page 3", align='R', new_x=XPos.LMARGIN, new_y=YPos.NEXT)
    pdf.set_draw_color(225, 228, 232)
    pdf.line(10, pdf.get_y(), 200, pdf.get_y())

    for repo_id, repo_folder in repos:
        page_no = section_start_pages[repo_id] + 2  # +1 cover, +1 TOC/overview shift
        pdf.set_font(fn, '', 11)
        pdf.cell(150, 8, f"Repository {repo_id}: {repo_folder}")
        pdf.cell(30, 8, f"page {page_no}", align='R',
                  new_x=XPos.LMARGIN, new_y=YPos.NEXT)
        pdf.set_draw_color(225, 228, 232)
        pdf.line(10, pdf.get_y(), 200, pdf.get_y())

    # ---- Executive Overview page ----
    pdf.add_page()
    pdf.set_font(fn, 'B', 16)
    pdf.set_text_color(*NAVY)
    pdf.cell(0, 12, "Executive Overview", new_x=XPos.LMARGIN, new_y=YPos.NEXT)
    pdf.set_text_color(0, 0, 0)
    pdf.ln(2)

    pdf.set_font(fn, '', 10)
    pdf.multi_cell(0, 5.5,
        f"This report summarizes the classification of {grand_total:,} research dataset "
        f"projects collected from {n_repos} repositories. Each project was assigned a "
        f"PROJECT_TYPE (QDA_PROJECT, QD_PROJECT, OTHER_PROJECT, or NOT_A_PROJECT) based on "
        f"its file composition, and classified against the ISIC Rev.5 taxonomy at the Section "
        f"and Division level. The following pages present per-repository breakdowns.")
    pdf.ln(6)

    pdf.set_font(fn, 'B', 12)
    pdf.set_text_color(*NAVY)
    pdf.cell(0, 8, "Overall Project Type Distribution", new_x=XPos.LMARGIN, new_y=YPos.NEXT)
    pdf.set_text_color(0, 0, 0)
    pdf.ln(2)
    draw_overall_type_bar(pdf, fn, overall_type_totals)
    pdf.ln(8)

    pdf.set_font(fn, 'B', 12)
    pdf.set_text_color(*NAVY)
    pdf.cell(0, 8, "Repository Summary", new_x=XPos.LMARGIN, new_y=YPos.NEXT)
    pdf.set_text_color(0, 0, 0)
    pdf.ln(2)

    pdf.set_fill_color(*NAVY)
    pdf.set_text_color(255, 255, 255)
    pdf.set_font(fn, 'B', 8.5)
    widths = [42, 16, 16, 16, 16, 16, 68]
    headers = ["Repository", "Total", "QDA", "QD", "OTHER", "N/A", "Dominant Class"]
    for w, h in zip(widths, headers):
        pdf.cell(w, 7.5, h, border=0, fill=True, align='C')
    pdf.ln()
    pdf.set_text_color(0, 0, 0)

    pdf.set_font(fn, '', 8)
    for idx, (repo_id, repo_folder) in enumerate(repos):
        c.execute("SELECT type, COUNT(*) FROM PROJECTS WHERE repository_id=? GROUP BY type", (repo_id,))
        tc = dict(c.fetchall())
        tot = sum(tc.values())
        c.execute("""SELECT primary_class, COUNT(*) cnt FROM PROJECTS
                     WHERE repository_id=? GROUP BY primary_class ORDER BY cnt DESC LIMIT 1""",
                  (repo_id,))
        dom = c.fetchone()
        dom_label = dom[0] if dom else "N/A"

        fill = ROW_ALT if idx % 2 == 0 else (255, 255, 255)
        pdf.set_fill_color(*fill)
        pdf.cell(widths[0], 6.5, f"{repo_id}: {repo_folder}", border=0, fill=True)
        pdf.cell(widths[1], 6.5, str(tot), border=0, fill=True, align='C')
        pdf.cell(widths[2], 6.5, str(tc.get("QDA_PROJECT", 0)), border=0, fill=True, align='C')
        pdf.cell(widths[3], 6.5, str(tc.get("QD_PROJECT", 0)), border=0, fill=True, align='C')
        pdf.cell(widths[4], 6.5, str(tc.get("OTHER_PROJECT", 0)), border=0, fill=True, align='C')
        pdf.cell(widths[5], 6.5, str(tc.get("NOT_A_PROJECT", 0)), border=0, fill=True, align='C')
        pdf.cell(widths[6], 6.5, dom_label[:44], border=0, fill=True)
        pdf.ln()

    pdf.set_draw_color(200, 204, 210)
    pdf.line(10, pdf.get_y(), 200, pdf.get_y())

    # ---- Per-repository sections ----
    for repo_id, repo_folder in repos:
        pdf.add_page()
        _render_repo_section(pdf, fn, c, repo_id, repo_folder)

    pdf.output(pdf_path)
    conn.close()
    print(f"📄 PDF Report Generated successfully -> {pdf_path} "
          f"(total pages: {pdf.page_no()}) [Slide 30 Aligned - Vector Graphics]")

# =====================================================================
# MAIN
# =====================================================================
if __name__ == "__main__":
    print("\n" + "="*65)
    print(f" SQ26 PHASE 2 - REAL DATA COLLECTOR | Student: {STUDENT_NAME} ({STUDENT_ID})")
    print("="*65)

    base_dir = resolve_output_dir()
    db_path = os.path.join(base_dir, DB_NAME)
    excel_path = os.path.join(base_dir, "Phase2_Classifications.xlsx")
    pdf_path = os.path.join(base_dir, "Phase2_Final_Report.pdf")

    conn = init_db(db_path)
    print(f"✅ DB initialised: {db_path}\n")

    grand_total = 0

    grand_total += fetch_dataverse(conn, REPO_DATAVERSE_NO)
    grand_total += fetch_ada_html_and_insert(conn)

    print("\n ⚠️ Attempting uni-halle (HTML). If a browser window appears, please complete the 'I'm not a robot' check.")
    grand_total += fetch_uni_halle_html_and_insert(conn)

    conn.close()

    print(f"\n{'='*65}")
    print(f" TOTAL REAL PROJECTS COLLECTED: {grand_total}")
    print(f"{'='*65}")

    export_excel(db_path, excel_path)
    generate_pdf(db_path, pdf_path)

    print(f"""
╔{'═'*63}╗
║ DELIVERABLES READY ║
╠{'═'*63}╣
║ DB (Git/tag classification-results):        ║
║ {db_path:<57} ║
║ XLSX (moo.uni1.de):                         ║
║ {excel_path:<57} ║
║ PDF (moo.uni1.de):                          ║
║ {pdf_path:<57} ║
╚{'═'*63}╝
""")